#!/usr/bin/env python3
import openpyxl
import os
import sys
import json
from openpyxl.worksheet.worksheet import Worksheet

#TODO: 시나리오마다 바뀔 수 있음
nationLevelMap = {
    '황제':7,
    '왕':6,
    '공':5,
    '주목':4,
    '주자사':3,
    '군벌':2,
    '호족':1,
    '방랑군':0
}

xlsxpath = sys.argv[1]
print(xlsxpath)
wb = openpyxl.load_workbook(xlsxpath, read_only=True)

def parseConfig(configSheet: Worksheet):
    result = {}
    for row in configSheet.iter_rows(min_row=2, values_only=True):
        print(row)
        if len(row) < 4:
            continue
        varName = row[0]
        value = row[3]

        if value is None:
            continue

        if isinstance(value, float) and float(int(value)) == value:
            value = int(value)

        if isinstance(value, str):
            values = value.split('\n')
            if len(values) > 1:
                value = values

        varNames = varName.split('.')
        target = result
        for idx, name in enumerate(varNames):
            if idx + 1 == len(varNames):
                target[name] = value
                break
            if not name in target:
                target[name] = {}
            target = target[name]

    if 'useCharSetting' in result:
        if result['useCharSetting'] > 0:
            result['fiction'] = 0
        else:
            result['fiction'] = 1
        del result['useCharSetting']
    return result

def extractNationList(nationSheet: Worksheet):
    jsonNationList = []
    nationChiefInfo = {}

    for row in nationSheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        if len(row) == 8:
            name, color, gold, rice, desc, tech, nationType, nationLevel = row
            cities = ''
        else:
            name, color, gold, rice, desc, tech, nationType, nationLevel, cities = row[:9]

        nationChiefInfo[name] = {}
        if len(row) >= 10:
            chief0 = str(row[9])
            if chief0:
                nationChiefInfo[name][chief0] = 12
        if len(row) >= 11:
            chief1 = str(row[10])
            if chief1:
                nationChiefInfo[name][chief1] = 11


        cities = list(map(str.strip, row[8].split(',')))

        if not cities:
            nationLevel = '방랑군'

        if nationLevel.isdigit():
            nationLevel = int(nationLevel)
        elif nationLevel in nationLevelMap:
            nationLevel = nationLevelMap[nationLevel]
        else:
            nationLevel = 1

        gold = int(gold)
        rice = int(rice)
        tech = int(tech)

        jsonNationList.append([name, color, gold, rice, desc, tech, nationType, nationLevel, cities])

    return jsonNationList, nationChiefInfo


def extractGeneralList(generalSheet: Worksheet, nationList={}, nationChiefInfo={}):
    nationInv = {'재야':0}
    for idx, nation in enumerate(nationList, 1):
        nationInv[nation[0]] = idx

    json_general_list = []
    names = {}
    for row_idx, row in enumerate(generalSheet.iter_rows(min_row=2, values_only= True)):
        if len(row) < 10:
            continue


        if len(row) < 13:
            row = list(row) + [None] * (13 - len(row))
        else:
            row = list(row[:13])

        #상성, 장수명, 전콘, 국가명, 도시, 통솔, 무력, 지력, 생년, 몰년, 성격, 고유특기, 고유대사
        declrow = [0, '', ('', str, int), ('', str, int), '', 10, 10, 10, 0, 0, '', '', '']
        for i, (raw_value, decl_type) in enumerate(zip(row, declrow)):
            if decl_type is None:
                continue
            if type(decl_type) is str:
                row[i] = str(raw_value or decl_type).strip()
                continue
            if type(decl_type) is int:
                if raw_value is None:
                    row[i] = 0
                elif type(raw_value) is int:
                    row[i] = raw_value
                elif type(raw_value) is str and raw_value.isdigit():
                    row[i] = int(raw_value)
                else:
                    raise RuntimeError('%d행 %d열 값이 숫자가 아닙니다: %s'%(row_idx+1, i+1, raw_value))
                continue

            if type(decl_type) is not tuple:
                raise RuntimeError('%d행 %d열 값이 잘못되었습니다: %s'%(row_idx+1, i+1, raw_value))

            if raw_value is not None and type(raw_value) not in decl_type:
                raise RuntimeError('%d행 %d열 값 타입이 이상합니다: %s'%(row_idx+1, i+1, raw_value))
            if raw_value is None:
                row[i] = decl_type[0]

            #상세 타입을 지정해야하지만... 귀찮다...
            if type(raw_value) is str:
                row[i] = raw_value.strip()
                continue

        상성: int
        장수명: str
        전콘: int or str
        국가명: int or str
        도시: str
        통솔: int
        무력: int
        지력: int
        생년: int
        몰년: int
        성격: str
        고유특기: str
        고유대사: str
        상성, 장수명, 전콘, 국가명, 도시, 통솔, 무력, 지력, 생년, 몰년, 성격, 고유특기, 고유대사 = row

        #이름
        장수명 = 장수명.strip()
        if 장수명 == '':
            continue
        print(장수명)
        #전콘
        if 전콘 == '':
            전콘 = None

        level = 0

        #국가
        국가명 = 국가명
        국가코드 = 0
        if 국가명 in nationInv:
            level = 1
            국가코드 = nationInv[국가명]
        else:
            국가코드 = 0
            if type(국가명) is str:
                if 국가명.isdigit():
                    국가코드 = int(국가명)
            elif type(국가명) is int:
                국가코드 = 국가명

            if 1 <= 국가코드 <= len(nationList):
                국가명 = nationList[국가코드 - 1][0]
                level = 1
            else:
                국가명 = ''
                level = 0

        if 도시 == '':
            도시 = None

        if level and 국가명 in nationChiefInfo:
            nationChiefDetail = nationChiefInfo[국가명]
            if 장수명 in nationChiefDetail:
                level = int(nationChiefDetail[장수명])

        if 성격 == '':
            성격 = None

        if 고유특기 == '':
            고유특기 = None

        if 고유대사 == '':
            고유대사 = None

        if 고유대사 is None:
            json_output = [상성, 장수명, 전콘, 국가코드, 도시, 통솔, 무력, 지력, level, 생년, 몰년, 성격, 고유특기]
        else:
            json_output = [상성, 장수명, 전콘, 국가코드, 도시, 통솔, 무력, 지력, level, 생년, 몰년, 성격, 고유특기, 고유대사]



        json_general_list.append(json_output)

        if 장수명 in names:
            raise RuntimeError('%s가 이미 있습니다!'%장수명)
        names[장수명] = 1
    return json_general_list, names


if '환경 변수' in wb:
    config = parseConfig(wb['환경 변수'])
    config['startYear'] -= 3
else:
    config = {
        'title':'타이틀'
    }

if '국가' in wb:
    nationInfo, nationChiefInfo = extractNationList(wb['국가'])
else:
    nationInfo = []
    nationChiefInfo = {}

generalList, names = extractGeneralList(wb['장수 목록'], nationInfo, nationChiefInfo)

with open('%s.json'%xlsxpath, 'wt', encoding='utf-8') as fp:
    fp.write(json.dumps(config, ensure_ascii=False, indent='    ')[:-2])
    fp.write(',\n')

    fp.write('    "nation":[\n        ')
    fp.write(',\n        '.join([json.dumps(nation, ensure_ascii=False) for nation in nationInfo]))
    fp.write('\n    ],\n')

    fp.write('    "diplomacy":[],\n')

    fp.write('    "general":[\n        ')
    fp.write(',\n        '.join([json.dumps(general, ensure_ascii=False) for general in generalList]))
    fp.write('\n    ]')

    names2 = []
    names3 = []
    if '확장 장수 목록' in wb.sheetnames:
        generalExList, names2 = extractGeneralList(wb['확장 장수 목록'], nationInfo, nationChiefInfo)

        for name in names2:
            if name in names:
                raise RuntimeError('%s가 일반 장수 및 확장 장수에 모두 있습니다!'%name)

        fp.write(',\n    "general_ex":[\n        ')
        fp.write(',\n        '.join([json.dumps(general, ensure_ascii=False) for general in generalExList]))
        fp.write('\n    ]')

    if '빙의 불가 장수 목록' in wb.sheetnames:
        generalNeutralList, names3 = extractGeneralList(wb['빙의 불가 장수 목록'], nationInfo, nationChiefInfo)

        for name in names3:
            if name in names:
                raise RuntimeError('%s가 일반 장수 및 빙의 불가 장수에 모두 있습니다!'%name)
        for name in names3:
            if name in names2:
                raise RuntimeError('%s가 확장 장수 및 빙의 불가 장수에 모두 있습니다!'%name)

        fp.write(',\n    "general_neutral":[\n        ')
        fp.write(',\n        '.join([json.dumps(general, ensure_ascii=False) for general in generalNeutralList]))
        fp.write('\n    ]')

    fp.write('\n}')